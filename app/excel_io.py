"""读取标准模板「正式题目」工作表，并导出查重题对。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from app.models import PairResult, Question

# 导出表中 B 侧列（编号B/题型B/题干B/选项B/原表行B）统一蓝字，与 A 侧黑字左右对照
_B_SIDE_COLS = (4, 6, 8, 10, 12)  # 1-based 列号
_BLUE_FONT = Font(color="0000FF")

SHEET_NAME = "正式题目"
REQUIRED_COLS = ("编号", "试题内容")
OPTION_COLS = ("候选项A", "候选项B", "候选项C", "候选项D", "候选项E", "候选项F")


class TemplateError(ValueError):
    """不是本工具要求的试题模板。"""


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_rows(path: Path) -> list[list]:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        import xlrd

        book = xlrd.open_workbook(str(path))
        if SHEET_NAME not in book.sheet_names():
            raise TemplateError(f"未找到工作表「{SHEET_NAME}」，请使用标准试题模板。")
        sheet = book.sheet_by_name(SHEET_NAME)
        return [
            [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        book = load_workbook(str(path), read_only=True, data_only=True)
        if SHEET_NAME not in book.sheetnames:
            book.close()
            raise TemplateError(f"未找到工作表「{SHEET_NAME}」，请使用标准试题模板。")
        sheet = book[SHEET_NAME]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        book.close()
        return rows
    raise TemplateError("仅支持 .xls / .xlsx 文件。")


def _find_header(rows: list[list]) -> tuple[int, dict[str, int]]:
    """在正式题目表前几行里定位表头，校验必填列。"""
    for idx, row in enumerate(rows[:20]):
        names = {_cell_str(c): i for i, c in enumerate(row) if _cell_str(c)}
        if all(col in names for col in REQUIRED_COLS):
            return idx, names
    raise TemplateError(
        f"工作表「{SHEET_NAME}」缺少表头列：{'、'.join(REQUIRED_COLS)}。"
    )


def load_questions(path: str | Path) -> list[Question]:
    """只读「正式题目」。空题干跳过。答案只读入供复核，不参与查重。"""
    path = Path(path)
    if not path.is_file():
        raise TemplateError("文件不存在。")
    rows = _read_rows(path)
    header_idx, col = _find_header(rows)
    questions: list[Question] = []
    for offset, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        def get(name: str) -> str:
            i = col.get(name)
            if i is None or i >= len(row):
                return ""
            return _cell_str(row[i])

        stem = get("试题内容")
        if not stem:
            continue
        code = get("编号") or f"行{offset}"
        options = [get(name) for name in OPTION_COLS]
        questions.append(
            Question(
                excel_row=offset,
                code=code,
                qtype=get("题型"),
                stem=stem,
                options=options,
                answer=get("答案"),
            )
        )
    if len(questions) < 2:
        raise TemplateError("有效试题不足 2 道，无法查重。")
    return questions


def export_pairs(
    dest: str | Path,
    questions: list[Question],
    pairs: list[PairResult],
    scores: list[float],
) -> None:
    """导出当前阈值下的题对（不是去重后的单题列表）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "查重结果"
    ws.append(
        [
            "序号",
            "相似度",
            "编号A",
            "编号B",
            "题型A",
            "题型B",
            "题干A",
            "题干B",
            "选项A",
            "选项B",
            "原表行A",
            "原表行B",
        ]
    )
    for seq, (pair, score) in enumerate(zip(pairs, scores), start=1):
        a, b = questions[pair.i], questions[pair.j]
        ws.append(
            [
                seq,
                round(float(score), 4),
                a.code,
                b.code,
                a.qtype,
                b.qtype,
                a.stem,
                b.stem,
                a.option_summary,
                b.option_summary,
                a.excel_row,
                b.excel_row,
            ]
        )
        row = ws.max_row
        for col in _B_SIDE_COLS:
            ws.cell(row=row, column=col).font = _BLUE_FONT
    # 表头同步：B 侧列名也标蓝
    for col in _B_SIDE_COLS:
        ws.cell(row=1, column=col).font = _BLUE_FONT
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
